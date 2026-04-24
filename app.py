from flask import Flask, request, jsonify, render_template, send_file, session, redirect
from config import Config
from models import db, StudentOMR, AnswerKey
from omr_detect import process_omr_file
from concurrent.futures import ProcessPoolExecutor, as_completed
import os
import uuid
import zipfile
import threading
import io
import pandas as pd
import pymysql
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from flask import send_file
import io

pymysql.install_as_MySQLdb()

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = "omr-secret"

db.init_app(app)

UPLOAD_FOLDER = app.config.get("UPLOAD_FOLDER", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
progress_store = {}

with app.app_context():
    db.create_all()


@app.route("/")
def home():
    return render_template("index.html")


def process_single_file(filepath, key_dict, batch_id):
    try:
        result = process_omr_file(filepath)

        if "error" in result:
            return None

        answers = result["answers"]
        level = result["level"].strip().lower()

        answers_json = {}
        verify_json = {}  
        final_answers = []

        total_score = 0
        correct_count = 0
        wrong_count = 0
        empty_count = 0

        for q, ans in answers.items():

            q = int(q)
            q_no = f"Q{str(q).zfill(2)}"

            correct = key_dict.get(("intermediate", q))

            ans = str(ans).strip().upper() if ans else None
            correct = str(correct).strip().upper() if correct else None

            is_correct = False

            if ans in [None, '-', '']:
                selected = "Empty"
                score = 0
                empty_count += 1

            elif "&" in str(ans):
                selected = "Multiple"
                score = 0
                wrong_count += 1

            else:
                selected = ans
                is_correct = (ans == correct)

                if is_correct:
                    correct_count += 1
                else:
                    wrong_count += 1

                if 1 <= q <= 25:
                    score = 2 if is_correct else 0

                elif 26 <= q <= 35:
                    score = 3 if is_correct else -1

                elif 36 <= q <= 40:
                    score = 4 if is_correct else -2

                else:
                    score = 0

            total_score += score
            answers_json[q_no] = selected
            
            verify_json[q_no] = {
                "selected": selected,
                "correct": correct,
                "is_correct": is_correct,
                "score": score
            }

            final_answers.append({
                "question": q_no,
                "value": selected,
                "correct_answer": correct,
                "is_correct": is_correct,
                "score": score
            })

        max_marks = (25 * 2) + (10 * 3) + (5 * 4)
        percentage = (total_score / max_marks) * 100 if max_marks else 0

        return {
            "db_data": {
                "name": result["name"],
                "level": level,
                "centre_number": result["centre_number"],
                "dob": result["dob"],
                "answers": answers_json,
                "verify_ans": verify_json,
                "score": round(total_score, 2),
                "batch_id": batch_id,
                "file_name": os.path.basename(filepath),
            },
            "row_data": {
                "name": result["name"],
                "centre_number": result["centre_number"],
                "level": level,
                "answers": final_answers,
                "dob":result["dob"],
                "total_score": total_score,
                "percentage": round(percentage, 2),
                "correct": correct_count,
                "wrong": wrong_count,
                "empty": empty_count
            }
        }

    except Exception as e:
        print("Worker Error:", e)
        return None


@app.route("/api/start", methods=["POST"])
def start_upload():
    batch_id = str(uuid.uuid4())

    progress_store[batch_id] = {
        "total": 0,
        "processed": 0,
        "status": "Starting",
        "results": []
    }

    session["latest_batch_id"] = batch_id

    return jsonify({"batch_id": batch_id})


@app.route("/api/upload", methods=["POST"])
def upload():

    batch_id = request.form.get("batch_id")

    if not batch_id:
        return jsonify({"error": "No batch_id"}), 400

    files = request.files.getlist("files")
    
    extracted_files = []
    
    for file in files:
        if file.filename.endswith(".zip"):
            zip_ref = zipfile.ZipFile(file.stream, 'r')

            for name in zip_ref.namelist():
                if name.lower().endswith((".jpg", ".jpeg", ".png")):
                    data = zip_ref.read(name)
                    original_name = os.path.basename(name)
                    file_name = os.path.splitext(original_name)[0]
                    ext = os.path.splitext(original_name)[1]
                    new_filename = f"{file_name}_{uuid.uuid4()}{ext}"
                    
                    temp_path = os.path.join(UPLOAD_FOLDER, new_filename)
                    with open(temp_path, "wb") as f:
                        f.write(data)
                    extracted_files.append(temp_path)
        else:
            original_name = os.path.basename(file.filename)
            file_name = os.path.splitext(original_name)[0]
            ext = os.path.splitext(original_name)[1]
            new_filename = f"{file_name}_{uuid.uuid4()}{ext}"
            
            path = os.path.join(UPLOAD_FOLDER, new_filename)
            file.save(path)
            extracted_files.append(path)

    progress_store[batch_id]["total"] = len(extracted_files)

    def background_process():

        with app.app_context():

            keys = AnswerKey.query.all()
            key_dict = {
                (k.level.strip().lower(), k.question_number): k.correct_answer
                for k in keys
            }

            files = extracted_files

            MAX_WORKERS = max(2, os.cpu_count() - 1)

            results_buffer = [None] * len(files)

            with ProcessPoolExecutor(max_workers=MAX_WORKERS) as executor:

                future_map = {
                    executor.submit(process_single_file, path, key_dict, batch_id): idx
                    for idx, path in enumerate(files)
                }

                for i, future in enumerate(as_completed(future_map)):

                    idx = future_map[future]
                    res = future.result()

                    if not res:
                        print("Worker returned None")
                        continue

                    student = StudentOMR(**res["db_data"])
                    db.session.add(student)
                    db.session.commit()

                    row = {
                        "key": str(student.id),
                        **res["row_data"]
                    }

                    results_buffer[idx] = row

                    progress_store[batch_id]["processed"] = i + 1
                    progress_store[batch_id]["status"] = f"Processing ({i+1}/{len(files)})"

            progress_store[batch_id]["results"] = [r for r in results_buffer if r]

            progress_store[batch_id]["status"] = "Completed"

    threading.Thread(target=background_process).start()

    return jsonify({"message": "Processing started"})


@app.route("/api/results/<batch_id>")
def get_results(batch_id):

    data = progress_store.get(batch_id)

    if not data:
        return jsonify({"results": []})

    offset = int(request.args.get("offset", 0))
    results = data["results"]

    new_results = results[offset:]

    total = data["total"]
    processed = data["processed"]
    percent = int((processed / total) * 100) if total else 0

    return jsonify({
        "results": new_results,
        "offset": offset + len(new_results),
        "total": total,
        "processed": processed,
        "percent": percent,
        "status": data["status"]
    })


@app.route("/api/export_latest")
def export_excel():

    batch_id = session.get("latest_batch_id")

    if not batch_id:
        return jsonify({"error": "No batch"}), 400

    students = StudentOMR.query.filter_by(batch_id=batch_id).order_by(StudentOMR.score.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    grey = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    headers = [
        "SL No", "File Name", "Name", "Centre Number",
        "Level", "DOB"
    ]
    
    for i in range(1, 41):
        headers.append(f"Q{str(i).zfill(2)}")

    headers += ["Correct", "Wrong", "Empty", "Score", "Percentage"]

    ws.append(headers)

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).fill = header_fill

    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        
    for idx, s in enumerate(students, start=1):

        answers = s.answers or {}
        verify = getattr(s, "verify_ans", {}) or {}

        correct_count = 0
        wrong_count = 0
        empty_count = 0

        row = [
            idx,
            getattr(s, "file_name", "-"),
            s.name or "-",
            s.centre_number or "-",
            s.level or "-",
            s.dob or "-"
        ]

        start_col = len(row) + 1

        for i in range(1, 41):
            q_key = f"Q{str(i).zfill(2)}"

            ans = answers.get(q_key, "-")
            v = verify.get(q_key, {})

            is_correct = v.get("is_correct", False)

            if ans in ["-", "Empty"]:
                empty_count += 1
            elif is_correct:
                correct_count += 1
            else:
                wrong_count += 1

            row.append(ans)

        score = s.score or 0
        max_marks = (25*2) + (10*3) + (5*4)
        percentage = round((score / max_marks * 100), 2) if max_marks else 0

        row += [correct_count, wrong_count, empty_count, score, percentage]

        ws.append(row)

        for i in range(40):
            col = start_col + i
            cell = ws.cell(row=idx + 1, column=col)

            q_key = f"Q{str(i+1).zfill(2)}"
            v = verify.get(q_key, {})

            if cell.value in ["-", "Empty"]:
                cell.fill = grey
            elif v.get("is_correct"):
                cell.fill = green
            else:
                cell.fill = red

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"OMR_Results_{timestamp}.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/save_answer_key", methods=["POST"])
def save_answer_key():

    data = request.json
    level = data.get("level", "").strip().lower()
    answers = data.get("answers", {})

    if not level:
        return jsonify({"error": "Level required"}), 400

    for q, ans in answers.items():

        q_no = int(q.replace("Q", ""))

        existing = AnswerKey.query.filter_by(
            level=level,
            question_number=q_no
        ).first()

        if existing:
            existing.correct_answer = ans
        else:
            db.session.add(AnswerKey(
                level=level,
                question_number=q_no,
                correct_answer=ans
            ))

    db.session.commit()

    return jsonify({"message": "Answer key saved successfully"})


@app.route("/api/get_answer_key/<level>")
def get_answer_key(level):

    keys = AnswerKey.query.filter_by(level=level).all()

    data = {
        f"Q{str(k.question_number).zfill(2)}": k.correct_answer
        for k in keys
    }

    return jsonify(data)

if __name__ == "__main__":
    app.run(debug=True)